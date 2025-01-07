import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

product_per_supplier = {}
total_value_per_supplier = {}
product_under_10_inv = {}

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)
    
    # Calculate the total quantity per supplier
    
    if supplier_name in product_per_supplier:
        current_quantity = product_per_supplier[supplier_name]
        product_per_supplier[supplier_name] = current_quantity + 1
    else:
        product_per_supplier[supplier_name] = 1
        
    # Calculate the total value per supplier
        
    if supplier_name in total_value_per_supplier:
        current_value = total_value_per_supplier[supplier_name]
        total_value_per_supplier[supplier_name] = current_value + (price * inventory)
    else:
        total_value_per_supplier[supplier_name] = price * inventory
        
    # Check if inventory is below 10
        
    if inventory < 10:
        product_under_10_inv[product_num] = inventory
        
    # Calculate the total value of the inventory
        
    inventory_price.value = inventory * price
        
print(product_per_supplier)
print(total_value_per_supplier)
print(product_under_10_inv)

inv_file.save("inventory_with_total_value.xlsx")